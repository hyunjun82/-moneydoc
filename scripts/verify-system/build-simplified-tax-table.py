# -*- coding: utf-8 -*-
"""
국세청 근로소득 간이세액표 엑셀 → 엔진용 JS 모듈 생성기

입력: moneydoc-data/sources/simplified-tax-table-2026-03-01.xlsx
      (홈택스 > 원천세 > 근로소득 간이세액표 > "2026.3.1. 이후" 엑셀 다운로드 원본, 수정 없음)
출력: lib/calc/tables/simplified-tax-2026-03.js  (손으로 고치지 말 것 — 이 스크립트로 재생성)

엑셀 구조 (2026-09-02 확인)
  시트 '근로소득간이세액표'
    5행~650행 : [이상(천원), 미만(천원), 가족1 … 가족11]   '-' = 0원
    651행     : ['10,000천원', None, 가족1 … 가족11]        월급여 정확히 1,000만원
    652행~    : 1,000만원 초과 구간 산식 (텍스트) → 아래 OVER 상수로 옮김 (원문 그대로 주석에 남김)
  시트 '소득령 별표2'
    비고 3 : 8~20세 자녀 조정 (1명 20,830 / 2명 45,830 / 3명↑ +33,330)
    비고 4 : 가족 11명 초과 → 11명 세액 − (10명 세액 − 11명 세액) × 초과 인원

사용: python scripts/verify-system/build-simplified-tax-table.py
"""
import json, re, sys, io
import openpyxl

SRC = 'moneydoc-data/sources/simplified-tax-table-2026-03-01.xlsx'
OUT = 'lib/calc/tables/simplified-tax-2026-03.js'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['근로소득간이세액표']

def cell_int(v):
    if v is None or v == '-': return 0
    if isinstance(v, (int, float)): return int(v)
    return int(str(v).replace(',', ''))

rows = []          # [from원, to원, [11]]
at10000 = None
over_text = []
for r in range(5, ws.max_row + 1):
    a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
    vals = [ws.cell(r, c).value for c in range(3, 14)]
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        rows.append([int(a) * 1000, int(b) * 1000, [cell_int(v) for v in vals]])
    elif isinstance(a, str) and a.strip() == '10,000천원':
        at10000 = [cell_int(v) for v in vals]
    elif isinstance(a, str):
        txt = ' '.join(str(x) for x in [a] + vals if x)
        over_text.append(txt.strip())

assert at10000 and len(at10000) == 11, 'at10000 누락'
assert rows[0][0] == 770000 and rows[-1][1] == 10000000, (rows[0], rows[-1])
# 연속성·단조성 검사 — 표가 깨졌으면 여기서 멈춘다
for i in range(1, len(rows)):
    assert rows[i][0] == rows[i-1][1], f'구간 끊김 {rows[i-1]} → {rows[i]}'
for fr, to, v in rows:
    for c in range(10):
        assert v[c] >= v[c+1], f'가족 수 증가에 세액 증가? {fr} {v}'

# 등간격 블록으로 압축 (5천/1만/2만원 단위)
blocks = []
for fr, to, v in rows:
    step = to - fr
    if blocks and blocks[-1]['step'] == step and blocks[-1]['end'] == fr:
        blocks[-1]['rows'].append(v); blocks[-1]['end'] = to
    else:
        blocks.append({'start': fr, 'end': to, 'step': step, 'rows': [v]})
for b in blocks:
    assert b['start'] + b['step'] * len(b['rows']) == b['end']

# 1,000만원 초과 산식 — 원문 텍스트에서 숫자를 뽑되, 아래 상수와 대조해 어긋나면 실패
OVER = [
    # (구간 상한 원, 가산 정액, 세율, 98% 적용 여부, 추가 정액)
    {'upTo': 14000000, 'fixed': 0,        'rate': 0.35, 'apply98': True,  'plus': 25000},
    {'upTo': 28000000, 'fixed': 1397000,  'rate': 0.38, 'apply98': True,  'plus': 0},
    {'upTo': 30000000, 'fixed': 6610600,  'rate': 0.40, 'apply98': True,  'plus': 0},
    {'upTo': 45000000, 'fixed': 7394600,  'rate': 0.40, 'apply98': False, 'plus': 0},
    {'upTo': 87000000, 'fixed': 13394600, 'rate': 0.42, 'apply98': False, 'plus': 0},
    {'upTo': None,     'fixed': 31034600, 'rate': 0.45, 'apply98': False, 'plus': 0},
]
formulas = [t for t in over_text if '(10,000천원인 경우의 해당 세액)' in t]
assert len(formulas) == 6, f'초과 산식 6개 기대, {len(formulas)}개'
for o, t in zip(OVER, formulas):
    nums = [int(x.replace(',', '')) for x in re.findall(r'\(([0-9,]+)원\)', t)]
    pct = [int(x) for x in re.findall(r'([0-9]+)% 상당액', t)]
    assert pct == [int(o['rate'] * 100 + 0.5)], (t, pct)
    assert ('98%' in t) == o['apply98'], t
    expect_nums = [n for n in [o['fixed'], o['plus']] if n]
    assert nums == expect_nums, (t, nums, expect_nums)

notes = wb['소득령 별표2']
note_text = '\n'.join(str(notes.cell(r, 1).value) for r in range(1, notes.max_row + 1) if notes.cell(r, 1).value)
assert '20,830원' in note_text and '45,830원' in note_text and '33,330원' in note_text, '자녀 조정 비고 변경?'

table = {
    'source': '국세청 홈택스 근로소득 간이세액표 (소득세법 시행령 별표2, 2026.3.1. 이후)',
    'file': SRC,
    'effectiveFrom': '2026-03-01',
    'minMonthly': 770000,
    'blocks': blocks,
    'at10000000': at10000,
    'over10000000': OVER,
    'childAdjust': {'one': 20830, 'two': 45830, 'perExtraOver2': 33330},
}

def js_rows(rs):
    return ',\n'.join('      [' + ','.join(str(x) for x in v) + ']' for v in rs)

blocks_js = ',\n'.join(
    '    { start: %d, end: %d, step: %d, rows: [\n%s\n    ] }' % (b['start'], b['end'], b['step'], js_rows(b['rows']))
    for b in blocks)
over_js = ',\n'.join(
    '    { upTo: %s, fixed: %d, rate: %s, apply98: %s, plus: %d }' % (
        'null' if o['upTo'] is None else o['upTo'], o['fixed'], o['rate'], 'true' if o['apply98'] else 'false', o['plus'])
    for o in OVER)

src = f'''/**
 * 근로소득 간이세액표 2026.3.1. 이후 — 국세청 홈택스 엑셀 원본을 그대로 옮긴 표.
 * 생성: python scripts/verify-system/build-simplified-tax-table.py  (손으로 고치지 말 것)
 * 원본: {SRC}
 *
 * 구조
 *   blocks       : 월급여 [start, end) 구간을 step 등간격으로 자른 행. rows[i][가족수-1] = 소득세(원, 100%)
 *                  77만원 미만은 표에 없음 = 0원. 1,000만원 미만까지 646행.
 *   at10000000   : 월급여 정확히 1,000만원일 때 가족수별 세액
 *   over10000000 : 1,000만원 초과 구간 산식 (엑셀 652~662행 원문)
 *                  세액 = at10000000[가족] + fixed + (월급여 − 구간하한) × (apply98 ? 0.98 : 1) × rate + plus
 *   childAdjust  : 8세 이상 20세 이하 자녀 조정 (별표2 비고 3). 음수면 0.
 *   가족 11명 초과 : 11명 세액 − (10명 세액 − 11명 세액) × 초과 인원 (별표2 비고 4)
 */
const SIMPLIFIED_TAX_2026_03 = {{
  source: {json.dumps(table['source'], ensure_ascii=False)},
  effectiveFrom: '2026-03-01',
  minMonthly: 770000,
  blocks: [
{blocks_js}
  ],
  at10000000: [{','.join(str(x) for x in at10000)}],
  over10000000: [
{over_js}
  ],
  childAdjust: {{ one: 20830, two: 45830, perExtraOver2: 33330 }},
}};

module.exports = {{ SIMPLIFIED_TAX_2026_03 }};
'''
with open(OUT, 'w', encoding='utf-8', newline='\n') as f:
    f.write(src)
n = sum(len(b['rows']) for b in blocks)
print(f'OK rows={n} blocks={[(b["start"],b["end"],b["step"],len(b["rows"])) for b in blocks]} at10000000={at10000[:3]}...')
print('over formulas (원문):'); [print('  ', t) for t in formulas]
